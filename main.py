from sys import stdout
from time import sleep
import numpy as np
import openpyxl
import threading
import copy
from random import random
from sys import setrecursionlimit
import matplotlib.pyplot as plt
setrecursionlimit(4000)

drop=5
max_step=1

max_depth=1000
min_delta=0.0005
sh=openpyxl.load_workbook("da.xlsx",read_only=1)["Sheet1"]
columns=[]
cnt=[]
for i in range(sh.max_column):
    cnt.append([])
def _float(inp):
    if inp==None:
        return None
    return float(inp)
def get_sh():
        result=[]
        for i in range(sh.max_column):
            result.append([])
        fr=tuple(sh.rows)
        for i in range(len(fr)):
            row=fr[i]
            #stdout.write(f"{i} {sh.max_row}\r")
            #inp=sh.cell(column=k+1,row=i+1).value
            lis=list(map(lambda x:_float(x.value),row))
            for i in range(len(lis)):
                if lis[i] != None:
                    result[i].append(lis[i])
        return result
def cost(inp):
    selected_y=inp
    selected_x=np.array(range(len(inp)))
    A = np.vstack([selected_x, np.ones(len(selected_x))]).T
    ((m, c),r) = np.linalg.lstsq(A,selected_y, rcond=None)[0:2]
    if r.size==0:
        return 0
    return r[0]
def three_division(interval,left,right):
    lline=left+(right-left)//3
    rline=left+((right-left)*2)//3
    L=rline-lline
    if L<=2*drop+2:
        return (left+right)//2
    lcost=cost(interval[:lline-drop])+cost(interval[lline+drop:])
    rcost=cost(interval[:rline-drop])+cost(interval[rline+drop:])
    if lcost<rcost:
        return three_division(interval,left,rline)
    else:
        return three_division(interval,lline,right)
def sign(inp):
    if inp>0:
        return 1
    if inp<0:
        return -1
    return 0
def find(interval,left,right,last_cost,k):
    global cnt
    cnt[k]+=1
    now_cost=0

    L=len(interval)
    next_lline=three_division(interval[:right],0,right)
    next_rline=three_division(interval[left:],0,len(interval)-left)
    if abs(next_lline-left)>max_step:
        lline=left+sign(next_lline-left)*max_step
    else:
        lline=next_lline
    if abs(right-next_rline)>max_step:
        rline=right+sign(next_rline-right)*max_step
    else:
        rline=next_rline
    if lline>rline:
        lline,rline=rline,lline
    if rline-lline<=2*drop:
        lline-=drop
        rline+=drop
    lcost=cost(interval[:lline-drop])
    rcost=cost(interval[lline+drop:rline-drop])+cost(interval[rline+drop:])
    now_cost=lcost+rcost
    if abs(last_cost-now_cost)/(last_cost)<min_delta:
        return (lline,rline)
    elif cnt[k]>max_depth:
        return (lline,rline)
    return find(interval,lline,rline,now_cost,k)
def main(k): 
    
    global cnt
    global plot_dic
    cnt[k]=0
    x = np.array(columns[2*k])
    y = np.array(columns[2*k+1])
    date_dic = []
    date_dic_2 = []
    R=[]
    l=10
    for i in range(1, len(x)-l+1):
        selected_x = x[i:i+l]
        selected_y = y[i:i+l]  # 确保 y 也被切片
        
        A = np.vstack([selected_x, np.ones(len(selected_x))]).T
        ((m, c),r) = np.linalg.lstsq(A, selected_y, rcond=None)[0:2]  # 使用切片后的 y 进行拟合
        r=1/((np.mean(selected_x*selected_y)/l-np.mean(selected_x)*np.mean(selected_y))/np.sqrt((np.mean(selected_x*selected_x)/l-np.mean(selected_x)*np.mean(selected_x))*(np.mean(selected_y*selected_y)/l-np.mean(selected_y)*np.mean(selected_y))))
        date_dic.append( m)
        date_dic_2.append(c)
        R.append(r)
        tmp_L=len(date_dic)
    (l,r)=find(date_dic,tmp_L//3,(2*tmp_L)//3,10,k)

    print(k,l,r,flush=1)
    plot_dic[k]=((0,l-drop),(l+drop,r-drop),(r+drop,len(x)-5))
    
columns=copy.deepcopy(get_sh())
plot_dic={}
threads={}
for k in range(openpyxl.load_workbook("da.xlsx",read_only=1)["Sheet1"].max_column//2):
    threads[k]=threading.Thread(target=main,args=(k,))
    threads[k].start()
    #main(k)
    #plt.scatter(x[:len(date_dic)],date_dic,1,edgecolors="red")
for k in threads:
    threads[k].join()
key=tuple(plot_dic.keys())
for figure in key:
    x = np.array(columns[2*figure])
    y = np.array(columns[2*figure+1])
    plt.scatter(x, y,1)
    for (m,n) in plot_dic[figure]:
        i=m
        j=n
        _x=x[i:j]
        _y=y[i:j]
        A = np.vstack([_x,np.ones(j-i)]).T
        (m, c) = np.linalg.lstsq(A,y[i:j], rcond=None)[0]
        r=1/((np.mean(_x*_y)/(j-i)-np.mean(_x)*np.mean(_y))/np.sqrt((np.mean(_x*_x)/(j-i)-np.mean(_x)*np.mean(_x))*(np.mean(_y*_y)/(j-i)-np.mean(_y)*np.mean(_y))))
        plt.plot(x, m*x + c, label=f'r={r} {i} to {j}:{m}x+{c}',color=(random(),random(),random(),0.8))
    plt.legend()
    plt.savefig(f"figs\\{figure}.png")
    plt.close()
